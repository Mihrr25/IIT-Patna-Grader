import openpyxl as op

def grader(workbook1):

    def col_letter(n):
        result = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            result = chr(65 + remainder) + result
    
        return result

    sheet = workbook1.active

    total_rows = sheet.max_row
    total_columns = sheet.max_column

    total_test=total_columns-2
    total_students=total_rows-3

    net=0
    exams={}
    for i in range(2,total_columns):
        column=col_letter(i+1)
        exams[sheet[column+'1'].value]={
        "max":sheet[column+'2'].value,
        "weight":sheet[column+"3"].value
        }
        net+=sheet[column+"3"].value

    if net!=100:
        return None
    
    workbook2 = op.Workbook()
    sheet2=workbook2.active

    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            target_cell = sheet2.cell(row=cell.row, column=cell.column, value=cell.value)
    
    form1="="
    for te in range(total_test):
        if(te):
            form1+=" +"
        co=col_letter(3+te)
        tem=f"{co}4/${co}$2*${co}$3"
        form1+=tem

    predata=[]
    for i in sheet2.iter_rows(values_only=True):
        predata.append(i)
    colplus=col_letter(total_columns+1)
    sheet2[colplus+'3']="Grand Total/100"
    sheet2.column_dimensions[colplus].width = 15
    data=[]
    for i in range(4,total_rows+1):
        new_formula = form1.replace("4", str(i))
        sheet2[colplus+str(i)]=new_formula
        valu=0
        for te in range(total_test):
            co=col_letter(3+te)
            valu+=sheet2[co+str(i)].value/sheet2[co+'2'].value*sheet2[co+'3'].value
        data.append([predata[i-1],valu])

    data = sorted(data, key=lambda x: x[1], reverse=True)
    for row_index, row_data in enumerate(data, start=4):
        for col_index, value in enumerate(row_data[0], start=1): 
            sheet2.cell(row=row_index, column=col_index, value=value)


    coltwo=col_letter(total_columns+2)
    sheet2[coltwo+'3']="Grand Total/100"
    sheet2.column_dimensions[coltwo].width = 15
    row=4
    aa=total_students/20
    aa=round(aa,0)
    while aa:
        sheet2[coltwo+str(row)]='AA'
        row+=1
        aa-=1
    ab=total_students*3/20
    ab=round(ab,0)
    while ab:
        sheet2[coltwo+str(row)]='AB'
        row+=1
        ab-=1
    bb=total_students/4
    bb=round(bb,0)
    while bb:
        sheet2[coltwo+str(row)]='BB'
        row+=1
        bb-=1
    bc=total_students*3/10
    bc=round(bc,0)
    while bc:
        sheet2[coltwo+str(row)]='BC'
        row+=1
        bc-=1
    cc=total_students*3/20
    cc=round(cc,0)
    while cc:
        sheet2[coltwo+str(row)]='CC'
        row+=1
        cc-=1
    cd=total_students/20
    cd=round(cd,0)
    while cd:
        sheet2[coltwo+str(row)]='CD'
        row+=1
        cd-=1
    while row<=total_rows:
        sheet2[coltwo+str(row)]='DD'
        row+=1
    
    sheet2.column_dimensions['A'].width = 13
    sheet2.column_dimensions['B'].width = 27

    colfour=col_letter(total_columns+4)
    colfive=col_letter(total_columns+5)
    colsix=col_letter(total_columns+6)
    colseven=col_letter(total_columns+7)
    coleight=col_letter(total_columns+8)

    sheet2[colfour+'1']="Total Students"
    sheet2.column_dimensions[colfour].width = 15
    sheet2[colfive+'1']=f"=COUNTA(A4:A{total_rows})"

    sheet2[colfour+'3']='grade'
    sheet2[colfive+'3']='old iapc reco'
    sheet2[colsix+'3']='Counts'
    sheet2[colseven+'3']='Round'
    sheet2[coleight+'3']='count verified'
    sheet2.column_dimensions[colfive].width = 15
    sheet2.column_dimensions[coleight].width = 15


    sheet2[colfour+'4']='AA'
    sheet2[colfour+'5']='AB'
    sheet2[colfour+'6']='BB'
    sheet2[colfour+'7']='BC'
    sheet2[colfour+'8']='CC'
    sheet2[colfour+'9']='CD'
    sheet2[colfour+'10']='DD'
    sheet2[colfour+'11']='F'
    sheet2[colfour+'12']='I'
    sheet2[colfour+'13']='PP'
    sheet2[colfour+'14']='NP'

    sheet2[colfive+'4']=5
    sheet2[colfive+'5']=15
    sheet2[colfive+'6']=25
    sheet2[colfive+'7']=30
    sheet2[colfive+'8']=15
    sheet2[colfive+'9']=5
    sheet2[colfive+'10']=5
    sheet2[colfive+'11']=0
    sheet2[colfive+'12']=0
    sheet2[colfive+'13']=0
    sheet2[colfive+'14']=0

    form2=f"={colfive}4%*{colfive}1"
    form3=f"=ROUND({colsix}4,0)"
    
    for i in range(4,15):
        sheet2[colsix+str(i)]=form2.replace('4',str(i))
        sheet2[colseven+str(i)]=form3.replace('4',str(i))
        sheet2[coleight+str(i)]=f"=COUNTIF({coltwo}4:{coltwo}{total_rows},{colfour}{str(i)})"
    sheet2.title="Grade Sorted" 

    sheet3=workbook2.copy_worksheet(sheet2)
    sheet3.title="Roll Sorted"
    second_data=[]
    for row in sheet3.iter_rows(min_row=4,values_only=True):
        temp=[]
        for i in range(len(row)):
            if(i==total_columns):
                temp.append(row[i+1])
                break
            temp.append(row[i])
        second_data.append(temp)
        # break

    second_data = sorted(second_data, key=lambda x: x[0])
    for row in range(4,total_rows+1):
        for col in range(1,total_columns+2):
            value=second_data[row-4][col-1]
            if col==total_columns+1:
                col+=1
            sheet3[col_letter(col)+str(row)]=value
    return workbook2
    # workbook2.save("op.xlsx")
